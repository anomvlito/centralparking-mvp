"""
excel.py — Central Parking MVP
Upload de Excel de operadores + reconciliación contra detecciones de cámara.

Endpoints:
  POST /api/excel/upload              — sube y parsea un .xlsx
  GET  /api/excel/reconcile?date=     — compara Excel vs cámara para una fecha
  GET  /api/excel/imports             — lista de archivos subidos
"""

import io
import re
import datetime
from typing import Optional
from zoneinfo import ZoneInfo

import openpyxl
from fastapi import APIRouter, UploadFile, File, HTTPException

from api.database import _db, now_cl

router = APIRouter()
_CL = ZoneInfo("America/Santiago")

EXCEL_COLS = {
    "SECTOR":                   "sector",
    "INGRESO":                  "ingreso",
    "SALIDA":                   "salida",
    "NOMBRE OPERADOR ENTRADA":  "operador_entrada",
    "RUT OPERADOR ENTRADA":     "rut_operador_entrada",
    "NOMBRE OPERADOR SALIDA":   "operador_salida",
    "RUT OPERADOR SALIDA":      "rut_operador_salida",
    "CALLE":                    "calle",
    "PATENTE":                  "plate",
    "MINUTOS TRANSCURRIDOS":    "minutos",
    "MINUTOS LIBRES":           "minutos_libres",
    "VALOR":                    "valor",
    "ESPACIOS":                 "espacios",
    "ESTADO":                   "estado",
}


def _parse_amount(raw) -> float:
    """'$8.000' → 8000.0  |  '$1.970' → 1970.0  |  None → 0"""
    if not raw:
        return 0.0
    cleaned = re.sub(r"[^\d,.]", "", str(raw)).replace(".", "").replace(",", ".")
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def _parse_dt(raw) -> Optional[datetime.datetime]:
    """'26-05-2026 06:47:46' → datetime"""
    if not raw:
        return None
    if isinstance(raw, datetime.datetime):
        return raw.replace(tzinfo=_CL) if raw.tzinfo is None else raw
    for fmt in ("%d-%m-%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S"):
        try:
            return datetime.datetime.strptime(str(raw).strip(), fmt).replace(tzinfo=_CL)
        except ValueError:
            continue
    return None


def _normalize_plate(plate: str) -> str:
    return re.sub(r"[\s\-]", "", str(plate)).upper()


# ─────────────────────── Upload ─────────────────────────────────────────────

@router.post("/api/excel/upload")
async def upload_excel(file: UploadFile = File(...)):
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Solo se aceptan archivos .xlsx o .xls")

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active

    # Detectar fila de encabezados (primera fila no vacía)
    headers = {}
    header_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if any(c for c in row):
            for j, cell in enumerate(row):
                if cell and str(cell).strip().upper() in EXCEL_COLS:
                    headers[str(cell).strip().upper()] = j
            if headers:
                header_row = i
                break

    if not headers or "PATENTE" not in headers or "INGRESO" not in headers:
        raise HTTPException(400, "Formato de Excel no reconocido — faltan columnas PATENTE e INGRESO")

    records = []
    date_min = date_max = None

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(row):
            continue
        def g(col):
            idx = headers.get(col)
            return row[idx] if idx is not None else None

        plate = _normalize_plate(g("PATENTE") or "")
        ingreso = _parse_dt(g("INGRESO"))
        if not plate or not ingreso:
            continue

        salida = _parse_dt(g("SALIDA"))
        valor  = _parse_amount(g("VALOR"))
        d = ingreso.date()
        if date_min is None or d < date_min:
            date_min = d
        if date_max is None or d > date_max:
            date_max = d

        records.append({
            "sector":               str(g("SECTOR") or ""),
            "calle":                str(g("CALLE") or ""),
            "plate":                plate,
            "ingreso":              ingreso,
            "salida":               salida,
            "operador_entrada":     str(g("NOMBRE OPERADOR ENTRADA") or ""),
            "rut_operador_entrada": str(g("RUT OPERADOR ENTRADA") or ""),
            "operador_salida":      str(g("NOMBRE OPERADOR SALIDA") or ""),
            "rut_operador_salida":  str(g("RUT OPERADOR SALIDA") or ""),
            "minutos":              int(g("MINUTOS TRANSCURRIDOS") or 0),
            "minutos_libres":       int(g("MINUTOS LIBRES") or 0),
            "valor":                valor,
            "espacios":             int(g("ESPACIOS") or 1),
            "estado":               str(g("ESTADO") or ""),
        })

    if not records:
        raise HTTPException(400, "El archivo no contiene registros válidos")

    with _db() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO excel_imports (filename, row_count, date_from, date_to)
                VALUES (%s, %s, %s, %s) RETURNING id
            """, (file.filename, len(records), date_min, date_max))
            import_id = cur.fetchone()["id"]

            for r in records:
                cur.execute("""
                    INSERT INTO excel_records
                        (import_id, sector, calle, plate, ingreso, salida,
                         operador_entrada, rut_operador_entrada,
                         operador_salida, rut_operador_salida,
                         minutos, minutos_libres, valor, espacios, estado)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (
                    import_id, r["sector"], r["calle"], r["plate"],
                    r["ingreso"], r["salida"],
                    r["operador_entrada"], r["rut_operador_entrada"],
                    r["operador_salida"],  r["rut_operador_salida"],
                    r["minutos"], r["minutos_libres"], r["valor"],
                    r["espacios"], r["estado"],
                ))

    return {
        "import_id":  import_id,
        "filename":   file.filename,
        "row_count":  len(records),
        "date_from":  str(date_min),
        "date_to":    str(date_max),
    }


# ─────────────────────── Reconciliación ─────────────────────────────────────

MATCH_WINDOW_MINUTES = 90   # ±90 min tolerancia entre Excel e ingreso de cámara

@router.get("/api/excel/reconcile")
async def reconcile(date: str, import_id: Optional[int] = None):
    """
    Compara registros del Excel contra detecciones de cámara para una fecha.
    Devuelve matched / camera_only (FRAUDE) / excel_only (cámara no vio).
    """
    try:
        target_date = datetime.date.fromisoformat(date)
    except ValueError:
        raise HTTPException(400, "Formato de fecha inválido — usar YYYY-MM-DD")

    day_start = datetime.datetime.combine(target_date, datetime.time.min, tzinfo=_CL)
    day_end   = datetime.datetime.combine(target_date, datetime.time.max, tzinfo=_CL)
    window    = datetime.timedelta(minutes=MATCH_WINDOW_MINUTES)

    with _db() as conn:
        with conn.cursor() as cur:
            # Detecciones de cámara del día (solo ENTRY)
            cur.execute("""
                SELECT id, plate, logged_at, confidence, image_path
                FROM detection_log
                WHERE action = 'ENTRY'
                  AND logged_at BETWEEN %s AND %s
                ORDER BY logged_at
            """, (day_start, day_end))
            camera_rows = cur.fetchall()

            # Registros del Excel para ese día
            query = """
                SELECT er.id, er.plate, er.ingreso, er.salida,
                       er.operador_entrada, er.valor, er.estado,
                       ei.filename
                FROM excel_records er
                JOIN excel_imports ei ON ei.id = er.import_id
                WHERE er.ingreso BETWEEN %s AND %s
            """
            params = [day_start, day_end]
            if import_id:
                query += " AND er.import_id = %s"
                params.append(import_id)
            query += " ORDER BY er.ingreso"
            cur.execute(query, params)
            excel_rows = cur.fetchall()

    # ── Matching ─────────────────────────────────────────────────────────────
    camera_list = [dict(r) for r in camera_rows]
    excel_list  = [dict(r) for r in excel_rows]

    matched      = []
    camera_only  = []
    excel_only   = []

    used_camera_ids = set()
    used_excel_ids  = set()

    for cam in camera_list:
        cam_plate = _normalize_plate(cam["plate"])
        cam_time  = cam["logged_at"]
        best_excel = None
        best_diff  = None

        for ex in excel_list:
            if ex["id"] in used_excel_ids:
                continue
            if _normalize_plate(ex["plate"]) != cam_plate:
                continue
            diff = abs((cam_time - ex["ingreso"]).total_seconds()) / 60
            if diff <= MATCH_WINDOW_MINUTES:
                if best_diff is None or diff < best_diff:
                    best_excel = ex
                    best_diff  = diff

        if best_excel:
            used_camera_ids.add(cam["id"])
            used_excel_ids.add(best_excel["id"])
            matched.append({
                "plate":              cam["plate"],
                "camera_time":        cam_time.strftime("%H:%M:%S"),
                "excel_ingreso":      best_excel["ingreso"].strftime("%H:%M:%S"),
                "diff_minutes":       round(best_diff, 1),
                "confidence":         float(cam["confidence"]),
                "valor":              float(best_excel["valor"]) if best_excel["valor"] else 0,
                "operador":           best_excel["operador_entrada"],
                "estado":             best_excel["estado"],
                "image_url":          f"/api/monitor/file/{cam['image_path']}" if cam["image_path"] else None,
            })
        else:
            camera_only.append({
                "plate":      cam["plate"],
                "camera_time": cam_time.strftime("%H:%M:%S"),
                "confidence": float(cam["confidence"]),
                "image_url":  f"/api/monitor/file/{cam['image_path']}" if cam["image_path"] else None,
            })

    for ex in excel_list:
        if ex["id"] not in used_excel_ids:
            excel_only.append({
                "plate":       ex["plate"],
                "excel_ingreso": ex["ingreso"].strftime("%H:%M:%S"),
                "excel_salida":  ex["salida"].strftime("%H:%M:%S") if ex["salida"] else None,
                "valor":         float(ex["valor"]) if ex["valor"] else 0,
                "operador":      ex["operador_entrada"],
                "estado":        ex["estado"],
            })

    excel_revenue  = sum(r["valor"]  for r in matched) + sum(r["valor"] for r in excel_only)
    camera_revenue = 0  # la cámara no registra tarifas

    return {
        "date": date,
        "summary": {
            "camera_total":  len(camera_list),
            "excel_total":   len(excel_list),
            "matched":       len(matched),
            "camera_only":   len(camera_only),
            "excel_only":    len(excel_only),
            "excel_revenue": excel_revenue,
        },
        "camera_only": camera_only,
        "matched":     matched,
        "excel_only":  excel_only,
    }


@router.get("/api/excel/imports")
async def list_imports():
    with _db() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT id, filename, row_count, date_from, date_to, imported_at
                FROM excel_imports
                ORDER BY imported_at DESC
                LIMIT 20
            """)
            rows = cur.fetchall()
    return [
        {
            "id":          r["id"],
            "filename":    r["filename"],
            "row_count":   r["row_count"],
            "date_from":   str(r["date_from"]),
            "date_to":     str(r["date_to"]),
            "imported_at": r["imported_at"].strftime("%Y-%m-%d %H:%M"),
        }
        for r in rows
    ]
